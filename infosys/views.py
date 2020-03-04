from django.shortcuts import render, redirect
from django.http import HttpResponse
from .models import Student, Class, Dormitory, Live, Event, Representative
from .forms import StudentForm, EventForm
from django.views.generic import ListView, DetailView, CreateView
from django.db.models import Q

from openpyxl import load_workbook, worksheet
import json


# class StudentListView(ListView):
#     model = Student
#     template_name = "infosys/inquir/students/list.html"
#     context_object_name = 'students'

#     def get(self, request):

def help(request):
    return render(request, 'help.html')


def student_list(request):
    search = request.GET.get('search')
    if search:
        student_list = Student.objects.filter(
            Q(student_name=search)|
            Q(student_id=search)|
            Q(student_tel=search)|
            Q(student_qq=search)|
            Q(student_address__contains=search)
        )
    else:
        search = ''
        student_list = Student.objects.all()
    context = {'students':student_list}
    return render(request, 'infosys/inquir/list/student.html', context)



class StudentDetailView(CreateView):
    def get(self, request, *args, **kwargs):
        student = Student.objects.get(student_id=kwargs['pk'])
        print(student)
        live = Live.objects.filter(student_id=student)
        print(live)
        if live:
            live = live[0]
        events = Event.objects.filter(student=student)
        context = {'student':student, 'events':events,'live':live}
        return render(request, "infosys/inquir/detail/student.html", context)
    def post(self, request, *args, **kwargs):
        form = EventForm(data=request.POST)
        if form.is_valid():
            event = form.save(commit=False)
            event.student = Student.objects.get(student_id=kwargs['pk'])
            event.save()
        return redirect('infosys:student_detail',pk=kwargs['pk'])

            
        


# class StudentDetailView(DetailView):
#     model = Student
#     template_name = "infosys/inquir/detail/student.html"
#     context_object_name = "student"

#     def get_object(self, queryset=None):
#         students = super().get_object(queryset=None)
#         return students


# class StudentCreateView(CreateView):
#     model = Student
#     template_name = "infosys/inquir/students/add.html"
#     fields = ['student_id', 'student_name', 'student_sex', 'student_birth', 'student_address', 'student_tel', 'student_qq', 'dormitory_id', 'class_id']
    
    


def student_add(request):
    classes = Class.objects.all()
    dormitory = Dormitory.objects.all()
    if request.method == 'POST':
        print(request.POST['bed_num'])
        student_info = StudentForm(data=request.POST)
        if student_info.is_valid():
            if request.POST['dormitory_id'] == '':
                student_info.save()
                return redirect('infosys:student_list')
            else:
                live = Live()
                student_info.save()
                live.student_id = request.POST['student_id']
                live.dormitory_id = request.POST['dormitory_id']
                live.bed_num = request.POST['bed_num']
                live.live_id = request.POST['dormitory_id']+request.POST['bed_num']
                live.save()
                return redirect('infosys:student_list')
        else:
            ErrorDict=student_info.errors
            Error_Str=json.dumps(ErrorDict)
            Error_Dict=json.loads(Error_Str)
            return HttpResponse(Error_Dict.values())
    else:
        student_info = StudentForm()
        context = {'student_info': student_info, 'class': classes, 'dormitory': dormitory}
        return render(request, 'infosys/import/students/add.html', context)

def delete_student(request, id):
    student = Student.objects.get(student_id=id)
    student.delete()
    return redirect("infosys:student_list")

def student_upload(request):
    if request.method == 'POST':
        excel = request.FILES.get('file')


        wb = load_workbook(filename=excel)
        ws = wb['Sheet1']
        list_key = []
        json = []

        row = ws.max_row
        column = ws.max_column

        for col in range(1, column+1):
            list_key.append(ws.cell(row=3, column=col).value)
        for row in range(4, row+1):
            dict_v = {}
            for col in range(1, column+1):
                values = ws.cell(row=row, column=col).value
                dict_v[list_key[col-1]] = str(values)
            json.append(dict_v)
        for single in json:
            info = Student()
            grade = single['class_id'][0:4]
            live_id = single['dormitory_id'] + single['bed_num']
            Class.objects.get_or_create(class_id=single['class_id'], grade=grade)
            Dormitory.objects.get_or_create(dormitory_id=single['dormitory_id'])
            info.student_tel = single['student_tel']
            info.student_sex = single['student_sex']
            info.student_qq = single['student_qq']
            info.student_name = single['student_name']
            info.student_id = single['student_id']
            info.student_birth = single['student_birth']
            info.student_address = single['student_address']
            info.class_id = Class(class_id=single['class_id'])
            # info.live_id = Live(live_id=live_id)
            info.save()
            Live.objects.get_or_create(live_id=live_id,bed_num=single['bed_num'], student_id=Student(student_id=single['student_id']), dormitory_id=Dormitory(dormitory_id=single['dormitory_id']))
            print(single['student_id'])
        return redirect('infosys:student_upload')
    else:
        return render(request, 'infosys/import/students/upload.html',)


#Class views here
def class_upload(request):
    if request.method == 'POST':
        excel = request.FILES.get('file')


        wb = load_workbook(filename=excel)
        ws = wb['Sheet1']
        list_key = []
        json = []

        row = ws.max_row
        column = ws.max_column

        for col in range(1, column+1):
            list_key.append(ws.cell(row=3, column=col).value)
        for row in range(4, row+1):
            dict_v = {}
            for col in range(1, column+1):
                values = ws.cell(row=row, column=col).value
                dict_v[list_key[col-1]] = str(values)
            json.append(dict_v)
        for single in json:
            info = Class()
            grade = single['class_id'][0:4]
            info.class_id = single['class_id']
            info.class_name = single['class_name']
            info.grade = grade
            info.major = single['major']
            info.college = single['college']
            info.save()
        return redirect('infosys:student_upload')
    else:
        return render(request, 'infosys/import/classes/upload.html',)


#Class views here

class ClassListView(ListView):
    model = Class
    template_name = "infosys/inquir/list/class.html"
    context_object_name = 'classes'


# class ClassDetailView(DetailView):
#     model = Class
#     template_name = "infosys/classes/detail.html"
#     context_object_name = "class"
#     def get_queryset(self):
#         class_id = self.request.POST.get('pk')
#         print (class_id)
#         return Student.objects.filter(class_id=class_id)

def class_detail(request, pk):
    students = Student.objects.filter(class_id=pk)
    detail = Class.objects.get(class_id=pk)
    context = {'class':detail, 'students':students}
    return render(request, "infosys/inquir/detail/class.html", context)


    
# Dormotory views here
class DormitoryListView(ListView):
    model = Dormitory
    template_name = "infosys/inquir/list/dormitory.html"
    context_object_name = 'dormitories'


# def dormitory_list(request):
#     dormitory_list = Dormitory.objects.all()
#     context = {'dormitories':dormitory_list}
#     return render(request, "infosys/dormitories/list.html", context)

# class DormitoryDetailView(DetailView):
#     model = Dormitory
#     template_name = "infosys/dormitories/detail.html"
#     context_object_name = "dormitory"

    # def get_context_data(self, **kwargs):
    #     dormitory = super().get_context_data(**kwargs)
    #     return dormitory

def dormitory_detail(request, pk):
    live = Live.objects.filter(dormitory_id=pk)
    # students = Student.objects.filter(live_id=live)
    detail = Dormitory.objects.get(dormitory_id=pk)
    context = {'dormitory':detail, 'students':live}
    return render(request, "infosys/inquir/detail/dormitory.html", context)