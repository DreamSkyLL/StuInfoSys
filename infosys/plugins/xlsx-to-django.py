from openpyxl import load_workbook, worksheet

wb = load_workbook(filename='data.xlsx', read_only=True)
ws = wb['Sheet1']
list_key = []
json = []

row = ws.max_row
column = ws.max_column
for col in range(1, column+1):
    print(col)
    list_key.append(ws.cell(row=3, column=col).value)
    print(list_key)
for row in range(4, row):
    dict_v = {}
    for col in range(1, column+1):
        values = ws.cell(row=row, column=col).value
        dict_v[list_key[col-1]] = [str(values)]
    json.append(dict_v)
for single in json:
    print(single)
a = ws.cell(row=3, column=10).value
print("列数：%d" %column)
print(a)